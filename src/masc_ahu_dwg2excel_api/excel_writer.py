from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .dxf_reader import TableData


def sanitize_sheet_title(title: str) -> str:
    safe_title = "".join(c for c in (title or "") if c not in r'[]:*?/\\').strip()
    return safe_title[:31] if len(safe_title) > 31 else safe_title


def dedupe_sheet_title(title: str, used: set[str], fallback: str) -> str:
    safe_title = sanitize_sheet_title(title) or fallback
    base = safe_title
    suffix = 1
    while safe_title in used:
        suffix += 1
        candidate = f"{base}_{suffix}"
        safe_title = candidate[:31]
    used.add(safe_title)
    return safe_title


def _estimate_text_width(value: str) -> int:
    """Approximate width by the longest line length (Excel width unit ~ character width)."""
    if value is None:
        return 0
    text = str(value)
    if not text:
        return 0
    return max(len(line) for line in text.splitlines() or [""])


def tables_to_workbook(
    tables: Iterable[TableData],
    *,
    autosize: bool = True,
    min_col_width: float = 8.0,
    max_col_width: float = 60.0,
    base_row_height: float = 15.0,
) -> Workbook:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    used_titles: set[str] = set()
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    wrap_alignment = Alignment(wrap_text=True)

    for index, table in enumerate(tables, start=1):
        safe_title = dedupe_sheet_title(table.name or "", used_titles, fallback=f"Table{index}")

        ws = wb.create_sheet(title=safe_title)

        col_max_chars: list[int] = []
        row_max_lines: list[int] = []

        for r_idx, row in enumerate(table.rows, start=1):
            while len(row_max_lines) < r_idx:
                row_max_lines.append(1)
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if autosize:
                    cell.alignment = wrap_alignment

                while len(col_max_chars) < c_idx:
                    col_max_chars.append(0)
                col_max_chars[c_idx - 1] = max(col_max_chars[c_idx - 1], _estimate_text_width(value))

                text = "" if value is None else str(value)
                line_count = max(1, text.count("\n") + 1) if text else 1
                row_max_lines[r_idx - 1] = max(row_max_lines[r_idx - 1], line_count)

        # Apply merged cells (convert 0-based to 1-based)
        for r1, c1, r2, c2 in getattr(table, "merges", []) or []:
            if r1 == r2 and c1 == c2:
                continue
            ws.merge_cells(
                start_row=r1 + 1,
                start_column=c1 + 1,
                end_row=r2 + 1,
                end_column=c2 + 1,
            )
            # Set border on every cell in the merged region for stable Excel rendering
            for rr in range(r1 + 1, r2 + 2):
                for cc in range(c1 + 1, c2 + 2):
                    mcell = ws.cell(row=rr, column=cc)
                    mcell.border = thin_border
                    if autosize:
                        mcell.alignment = wrap_alignment
            cell = ws.cell(row=r1 + 1, column=c1 + 1)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if autosize:
            # Column width: estimated from longest text per column
            for c_idx, max_chars in enumerate(col_max_chars, start=1):
                # +2 padding; 1.2x factor for CJK / mixed content
                width = float(max_chars) * 1.2 + 2.0
                width = max(min_col_width, min(max_col_width, width))
                ws.column_dimensions[get_column_letter(c_idx)].width = width

            # Row height: estimated from line-break count (only meaningful with wrap_text)
            for r_idx, max_lines in enumerate(row_max_lines, start=1):
                if max_lines > 1:
                    ws.row_dimensions[r_idx].height = base_row_height * float(max_lines)

    # Keep at least one sheet to avoid save failure on empty workbook
    if not wb.sheetnames:
        wb.create_sheet("Empty")

    return wb


def save_workbook_for_dxf(output_dir: Path, dxf_path: Path, wb: Workbook, overwrite: bool = False) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    excel_name = dxf_path.with_suffix(".xlsx").name
    excel_path = output_dir / excel_name

    if excel_path.exists() and not overwrite:
        raise FileExistsError(f"Output file already exists: {excel_path}")

    wb.save(excel_path)
    return excel_path
