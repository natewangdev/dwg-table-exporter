from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment

from .dxf_reader import TableData


def tables_to_workbook(tables: Iterable[TableData]) -> Workbook:
    wb = Workbook()
    # 删除默认 sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    used_titles: set[str] = set()

    for index, table in enumerate(tables, start=1):
        title = table.name or f"Table{index}"
        # Excel sheet 名长度和字符有限制，这里做一个简单清洗
        safe_title = "".join(c for c in title if c not in r'[]:*?/\\').strip()
        if not safe_title:
            safe_title = f"Table{index}"
        if len(safe_title) > 31:
            safe_title = safe_title[:31]

        base = safe_title
        suffix = 1
        while safe_title in used_titles:
            suffix += 1
            candidate = f"{base}_{suffix}"
            safe_title = candidate[:31]
        used_titles.add(safe_title)

        ws = wb.create_sheet(title=safe_title)
        for r_idx, row in enumerate(table.rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 应用合并单元格（0-based -> 1-based）
        for r1, c1, r2, c2 in getattr(table, "merges", []) or []:
            if r1 == r2 and c1 == c2:
                continue
            ws.merge_cells(
                start_row=r1 + 1,
                start_column=c1 + 1,
                end_row=r2 + 1,
                end_column=c2 + 1,
            )
            # 合并单元格的样式以左上角单元格为准，这里将其设置为水平/垂直居中
            cell = ws.cell(row=r1 + 1, column=c1 + 1)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 如果没有任何表格，保留一个空 sheet，避免保存失败
    if not wb.sheetnames:
        wb.create_sheet("Empty")

    return wb


def save_workbook_for_dxf(output_dir: Path, dxf_path: Path, wb: Workbook, overwrite: bool = False) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    excel_name = dxf_path.with_suffix(".xlsx").name
    excel_path = output_dir / excel_name

    if excel_path.exists() and not overwrite:
        raise FileExistsError(f"输出文件已存在: {excel_path}")

    wb.save(excel_path)
    return excel_path

